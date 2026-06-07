from io import BytesIO
from typing import Optional
from uuid import UUID

from openpyxl import Workbook, load_workbook
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.inventory import InventoryBatch, InventoryCategory, InventoryItem
from app.models.tenant_config import Branch
from app.schemas.common import PaginatedResponse, PaginationParams
from app.schemas.inventory import InventoryItemCreate, InventoryItemUpdate
from app.services.audit_service import AuditService


class InventoryService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.audit = AuditService(db)

    async def list_items(
        self, tenant_id: UUID, params: PaginationParams, branch_id: Optional[UUID] = None
    ) -> PaginatedResponse:
        query = select(InventoryItem).where(InventoryItem.tenant_id == tenant_id, InventoryItem.deleted_at.is_(None))
        if branch_id:
            query = query.where(InventoryItem.branch_id == branch_id)
        if params.search:
            term = f"%{params.search}%"
            query = query.where(or_(InventoryItem.name.ilike(term), InventoryItem.name_ar.ilike(term), InventoryItem.sku.ilike(term)))
        count_result = await self.db.execute(select(func.count()).select_from(query.subquery()))
        total = count_result.scalar() or 0
        sort_col = getattr(InventoryItem, params.sort_by or "name", InventoryItem.name)
        query = query.order_by(sort_col.desc() if params.sort_order == "desc" else sort_col.asc())
        query = query.offset((params.page - 1) * params.page_size).limit(params.page_size)
        result = await self.db.execute(query)
        items = list(result.scalars().all())

        enriched = []
        for item in items:
            qty_result = await self.db.execute(
                select(func.coalesce(func.sum(InventoryBatch.quantity), 0)).where(InventoryBatch.item_id == item.id)
            )
            total_qty = float(qty_result.scalar() or 0)
            item_dict = {c.name: getattr(item, c.name) for c in item.__table__.columns}
            item_dict["total_quantity"] = total_qty
            enriched.append(item_dict)

        pages = (total + params.page_size - 1) // params.page_size if params.page_size else 0
        return PaginatedResponse(items=enriched, total=total, page=params.page, page_size=params.page_size, pages=pages)

    async def get_item(self, tenant_id: UUID, item_id: UUID) -> Optional[InventoryItem]:
        result = await self.db.execute(
            select(InventoryItem).where(
                InventoryItem.id == item_id, InventoryItem.tenant_id == tenant_id, InventoryItem.deleted_at.is_(None)
            )
        )
        return result.scalar_one_or_none()

    async def create_item(self, tenant_id: UUID, data: InventoryItemCreate, user_id: UUID) -> InventoryItem:
        item = InventoryItem(tenant_id=tenant_id, **data.model_dump())
        self.db.add(item)
        await self.db.flush()
        await self.audit.log(
            tenant_id=tenant_id, user_id=user_id, action="create", module="inventory",
            entity_type="inventory_item", entity_id=str(item.id), new_values=data.model_dump(mode="json"),
        )
        return item

    async def update_item(
        self, tenant_id: UUID, item_id: UUID, data: InventoryItemUpdate, user_id: UUID
    ) -> Optional[InventoryItem]:
        item = await self.get_item(tenant_id, item_id)
        if not item:
            return None
        for key, value in data.model_dump(exclude_unset=True).items():
            setattr(item, key, value)
        await self.db.flush()
        await self.audit.log(
            tenant_id=tenant_id, user_id=user_id, action="update", module="inventory",
            entity_type="inventory_item", entity_id=str(item.id),
            new_values=data.model_dump(exclude_unset=True, mode="json"),
        )
        return item

    async def delete_item(self, tenant_id: UUID, item_id: UUID, user_id: UUID) -> bool:
        item = await self.get_item(tenant_id, item_id)
        if not item:
            return False
        item.deleted_at = func.now()
        await self.audit.log(
            tenant_id=tenant_id, user_id=user_id, action="delete", module="inventory",
            entity_type="inventory_item", entity_id=str(item.id),
        )
        return True

    @staticmethod
    def generate_import_template() -> bytes:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"
        headers = ["SKU", "Name", "Name_AR", "Category", "Unit", "Unit_Cost", "Quantity", "Reorder_Level", "Batch_Number"]
        ws.append(headers)
        ws.append(["GLV-001", "Nitrile Gloves", "قفازات", "glove", "box", 150, 100, 10, "B-001"])
        ws.append(["TUB-EDTA", "EDTA Tube", "أنبوب EDTA", "tube", "piece", 3.5, 500, 50, "B-002"])
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def import_from_excel(self, tenant_id: UUID, content: bytes, user_id: UUID, branch_id: Optional[UUID] = None) -> dict:
        if not branch_id:
            branch_id = await self.db.scalar(
                select(Branch.id).where(Branch.tenant_id == tenant_id, Branch.deleted_at.is_(None)).limit(1)
            )
        wb = load_workbook(BytesIO(content), read_only=True)
        ws = wb.active
        created, updated = 0, 0
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or not row[0]:
                continue
            sku, name, name_ar, category, unit, unit_cost, quantity, reorder, batch = (
                str(row[0]).strip(),
                str(row[1] or "").strip(),
                str(row[2] or name or "").strip(),
                str(row[3] or "other").strip().lower(),
                str(row[4] or "piece").strip(),
                float(row[5] or 0),
                float(row[6] or 0),
                float(row[7] or 0),
                str(row[8] or f"B-{row[0]}").strip(),
            )
            try:
                cat = InventoryCategory(category)
            except ValueError:
                cat = InventoryCategory.OTHER
            existing = await self.db.scalar(
                select(InventoryItem).where(
                    InventoryItem.tenant_id == tenant_id, InventoryItem.sku == sku, InventoryItem.deleted_at.is_(None)
                )
            )
            if existing:
                existing.name = name or existing.name
                existing.name_ar = name_ar or existing.name_ar
                existing.unit_cost = unit_cost
                existing.reorder_level = reorder
                updated += 1
                item = existing
            else:
                item = InventoryItem(
                    tenant_id=tenant_id, branch_id=branch_id, sku=sku, name=name, name_ar=name_ar,
                    category=cat, unit=unit, unit_cost=unit_cost, reorder_level=reorder,
                )
                self.db.add(item)
                await self.db.flush()
                created += 1
            if quantity > 0:
                self.db.add(
                    InventoryBatch(
                        tenant_id=tenant_id, item_id=item.id, branch_id=branch_id,
                        batch_number=batch, quantity=quantity, unit_cost=unit_cost,
                    )
                )
        await self.db.flush()
        return {"created": created, "updated": updated, "total_rows": len(rows)}
